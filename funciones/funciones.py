import openpyxl

## comparacion de color en la casilla

#casilla.fill.start_color.index
# AZUl = 10
# Naranja = 12
# Amarillo = FFFFFF00
# Rojo 15

def normalize(s):
    replacements = (
        ("á", "a"),
        ("é", "e"),
        ("í", "i"),
        ("ó", "o"),
        ("ú", "u"),
    )
    for a, b in replacements:
        s = s.replace(a, b).replace(a.upper(), b.upper())
    return s

## Rango de los nombres de los archivos
def rango(ws,r_initial,r_final):
    # Param: ws = hoja con la que estamos trabajando
    # Value: Variable activa donde asignemos el valor dado por la funcion
        #    openpyxl.load_workbook
    # Param: r_initial = celda donde comienzan los nombres
    # Value: String con el formato cordenada de excel, ejem A2
    # Param: r_final = celda donde terminan los nombres
    # Value: Strign con el formato cordenada de excel, ejemplo C4
    
    cell_range = ws[r_initial:r_final]
    return cell_range

def delete_format(cell_item):
    cell_item.value = cell_item.value.upper()
    cell_item.value = normalize(cell_item.value)
    cell_item.value = cell_item.value.replace(' ','')
    cell_item.value = cell_item.value.replace('1','')
    cell_item.value = cell_item.value.replace('2','')
    cell_item.value = cell_item.value.replace('3','')
    cell_item.value = cell_item.value.replace('4','')
    cell_item.value = cell_item.value.replace('5','')
    cell_item.value = cell_item.value.replace('6','')
    cell_item.value = cell_item.value.replace('7','')
    cell_item.value = cell_item.value.replace('8','')
    cell_item.value = cell_item.value.replace('9','')
    cell_item.value = cell_item.value.replace('0','')
    cell_item.value = cell_item.value.replace('.','')
    return cell_item

"""def primer_archivo():
    arr_comparacion = []
    for row_doc1 in wsa1.iter_rows(min_row=4, min_col=19, max_col=20, max_row=320):
        cambio = row_doc1[0]
        if cambio.fill.start_color.index == 9:
            cambio = delete_format(cambio)
            cambio = cambio.value.splitlines()
            if type(cambio) == list:
                for i in cambio:
                    arr_comparacion1.append(i)
            else:
                arr_comparacion1.append(cambio.value)"""
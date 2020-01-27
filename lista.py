import openpyxl
from funciones.funciones import rango,delete_format,normalize

doc1 = openpyxl.load_workbook('archivos/DOSIS_SEPTIEMBRE_2019.xlsx')
doc2 = openpyxl.load_workbook('archivos/FER-EXCEL.xlsx')
wsa1 = doc1.active
wsa2 = doc2.active

#idk = delete_format(wsa1['S4'])

arr_comparacion1 = []
arr_comparacion2 = []
for row_doc1 in wsa1.iter_rows(min_row=4, min_col=19, max_col=20, max_row=320):
    cambio = row_doc1[0]
    if cambio.fill.start_color.index == 9:
        cambio = delete_format(cambio)
        cambio = cambio.value.splitlines()
        if type(cambio) == list:
            for i in cambio:
                arr_comparacion1.append(i)
        else:
            arr_comparacion1.append(cambio.value)


for row_doc2 in wsa2.iter_rows(min_row=2, min_col=2, max_col=3, max_row=889):
    cambio = row_doc2[0]
    if cambio.fill.start_color.index == '00000000':
        cambio = delete_format(cambio)
        cambio = cambio.value.splitlines()
        if type(cambio) == list:
            for i in cambio:
                arr_comparacion2.append(i)
        else:
            arr_comparacion2.append(cambio.value)

print(len(arr_comparacion1))
print(len(arr_comparacion2))
## Borrar los siguientes dos # (uno por uno) para poder vizualizar
## los arreglos de ambos archivos purgados

#print(arr_comparacion1)
#print(arr_comparacion2)